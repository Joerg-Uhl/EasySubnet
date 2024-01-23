import sys, os, openpyxl, re, math, numpy as np, time  # openpyxl: Erstellen und Bearbeiten von Excel-Dateien
from openpyxl.styles import Font, Alignment
from zsys import addierer, dez_in_misc  # eigenes Modul mit Funktion für Umwandlung und Addition von Zahlen jeglicher Zahlensysteme


wd = os.getcwd()
programmablauf = f"""\n\n
Keine Daten-Datei in {wd} gefunden! Manuelle Dateneingabe wird gestartet.\n
Du kannst das nächste Mal auch eine Excel- oder Textdatei in der jetzt folgenden Form vorbereiten 
und im aktuellen Arbeitsverzeichnis ({wd}) unter dem Dateinamen "EasySubnetDaten.xlsx/txt" speichern.
Leere Zeilen sind nicht erlaubt!\n
Gib unten das folgende ein: 
<Netz-ID> <Subnetzmaske>
<Anzahl Hosts (ohne Netz-ID und Broadcast-Adresse)> [Subnetz-Name] [p/pp]
\np oder pp steht für Priorität wenn du bestimmte Subnetze doppelt bzw. viermal so groß wie das Minimum 
machen möchtest. Namen und Prioritäten sind optional. Die Reihenfolge der Angaben ab der 2. Zeile ist egal.
\nBeispiel:
192.168.0.0 255.255.254.0
Verwaltung 95 p
30
50 Marketing
pp 20
...

Der Output enthält eine Version mit minimal großen Subnetzen, eine Version mit automatischer Verdoppelung
falls IP-Adressen übrig sind (Subnetze mit geringster Anzahl freier IP-Adressen werden als erstes verdoppelt) 
und eine Version entsprechend der Priorisierung.\n
Wenn du fertigt bist drücke noch einmal die Eingabetaste.\n
Standardausgabe ist die Excel-Datei "output0.xlsx" im akutellen Arbeitsverzeichnis.
Falls Excel nicht verfügbar ist wird eine Text-Datei "output0.txt" erstellt."""


### Funktionen

def daten_einlesen():
    ### Excel-Datei wird versucht zu öffnen
    #global eingabe
    eingabe = ""
    try:
        wb = openpyxl.load_workbook("EasySubnetDaten.xlsx")
        sheet = wb.active
        anzahl_zeilen = sheet.max_row + 1
        eingabe = "excel"
    except:
        pass

    ### Text-Datei wird versucht zu öffnen, sonst im except-Block manuelle Eingabe
    if eingabe == "":
        try:
            with open("EasySubnetDaten.txt") as txt:
                zeilen = txt.readlines()
                eingabe = "text"
        except:
            eingabe = "manuell"
            print(programmablauf)
    l_hosts = []    # Liste für die Anzahl der Hosts in den Subnetzen
    l_prio = []     # Liste für die Prioritäten, falls gesetzt
    l_name = []     # Liste für die Namen der Subnetze, falls angegeben
    l_eingabe = []  # zuerst wird alles als 1/2/3er-Tupel/Liste in der Liste l_eingabe gesammelt und später l_hosts, l_prio, l_name zugeordnet
    if eingabe == "excel":
        netz_id_ges = sheet.cell(1, 1).value        # Netz-ID des Gesamtnetzes wird identifiziert
        subnetzmaske = sheet.cell(1, 2).value       # Subnetzmaske des Gesamtnetzes wird identifiziert
        for row in sheet.iter_rows(min_row=2, max_row=anzahl_zeilen, max_col=3, values_only=True):
            if row[0:3] == (None, None, None):      # trifft die Schleife auf eine leere Zeile, wird die Datenaufnahme beendet
                break
            l_eingabe.append(row)
    elif eingabe == "text":
        netz_id_ges, subnetzmaske = zeilen[0].split()[0], zeilen[0].split()[1]
        for zeile in zeilen[1:]:
            if zeile == "\n":                       # trifft die Schleife auf eine leere Zeile, wird die Datenaufnahme beendet
                break
            l_eingabe.append(zeile.split())
    elif eingabe == "manuell":
        val = input()
        netz_id_ges = val.split()[0]
        subnetzmaske = val.split()[1]
        while True:
            val = input()
            if val == "":                           # trifft die Schleife auf eine leere Zeile, wird die Datenaufnahme beendet
                break
            l_eingabe.append(val.split())
    pattern_hosts = re.compile(r"^[1-9][0-9]?[0-9]?[0-9]?[0-9]?$")  # anhand von Mustern werden die Daten in die Listen l_hosts,...
    pattern_names = re.compile(r"^None$|^p$|^pp$|^[1-9][0-9]?[0-9]?[0-9]?[0-9]?$")  # ...l_name,...
    pattern_prios = re.compile(r"^[p]{1,2}$")   # ...l_prio einsortiert
    for i, zeile in enumerate(l_eingabe):
        count_l_hosts = 0   # pro Durchlauf der inneren Schleife darf jede Liste um max. 1 Element erweitert werden
        count_l_prio = 0
        count_l_name = 0
        for item in zeile:
            if pattern_hosts.match(str(item)) and count_l_hosts == 0:
                l_hosts.append(item)
                count_l_hosts = 1
            elif pattern_prios.match(str(item)) and count_l_prio == 0:
                l_prio.append(item)
                count_l_prio = 1
            elif not pattern_names.match(str(item)) and count_l_name == 0:  # alles, was nicht None, p, pp, oder eine Zahl ist, könnte ein Name sein
                l_name.append(item)
                count_l_name = 1
        if len(l_hosts) <= i:   # falls in der inneren Schleife kein Element hinzugefügt wurde, dann füge hinzu...
            l_hosts.append(f"Kann Hostanzahl nicht erkennen: {zeile}")
        if len(l_name) <= i:
            l_name.append(i+1)  # bei fehlendem Namen bekommt das Subnetz eine Nummer
        if len(l_prio) <= i:
            l_prio.append("leer")
    return netz_id_ges, subnetzmaske, l_hosts, l_prio, l_name

def datencheck():
    l_fehler = []
    pattern_netz_id_ges = re.compile(r"^\d?\d?\d\.\d?\d?\d\.\d?\d?\d\.\d?\d?\d$")
    if not pattern_netz_id_ges.match(netz_id_ges):    # Netz-ID wird auf korrekte Struktur und korrekte Zeichen geprüft
        l_fehler.append(netz_id_ges)
    if not pattern_netz_id_ges.match(subnetzmaske):   # Subnetzmaske wird auf korrekte Struktur und korrekte Zeichen geprüft
        l_fehler.append(subnetzmaske)
    netz_id_ges_oktette = netz_id_ges.split(".")
    for okt in netz_id_ges_oktette:
        try:
            if not 0 <= int(okt) <= 255:    # liegen alle Oktette der Netz-ID zwichen 0 und 255?
                l_fehler.append(netz_id_ges)
                break
        except:
            print("Laufzeitfehler: Umwandlung der Oktette der Netz-ID in Integer-Datentyp misslungen")
            input("Zum Beenden des Programms beliebige Taste drücken!")
            sys.exit()
    subnetzmaske_oktette = subnetzmaske.split(".")
    for i, okt in enumerate(subnetzmaske_oktette):   # sind alle Oktette der Subnetzmaske zulässig?
        try:
            if not int(okt) in (0,128,192,224,240,248,252,254,255):
                l_fehler.append(subnetzmaske)
                break
        except:
            print("Laufzeitfehler: Umwandlung der Oktette der Subnetzmaske in Integer-Datentyp misslungen")
            input("Zum Beenden des Programms beliebige Taste drücken!")
            sys.exit()
        try:
            if int(okt) < 255 and i < 3:    # Wenn ein Oktett < 255 ist, sind alle restlichen Oktette 0?
                if not "".join(subnetzmaske_oktette[i+1:]) == "0"*(3-i):
                    l_fehler.append(subnetzmaske)
                    break
        except:
            print("Laufzeitfehler: Umwandlung der Oktette der Subnetzmaske in Integer-Datentyp misslungen")
            input("Zum Beenden des Programms beliebige Taste drücken!")
            sys.exit()
    if len(l_hosts) == 0:   # Fehler, wenn auf Netz-ID und Subnetzmaske keine weiteren Angaben erfolgen
        l_fehler.append("Keine Hostanzahl gefunden")
    for host in l_hosts:    # Fehler, wenn keine Zahl erkannt werden konnte
        if "Kann Hostanzahl nicht erkennen" in host:
            l_fehler.append(host)
    return l_fehler

def check_genug_IPs(l_geplante_IP_Adressen):
    ### prüft, ob die IP-Range des Gesamtnetzes überhaupt groß genug ist für alle Hosts
    subnetzmaske_bin = ""
    for okt in subnetzmaske.split("."):     # Subnetzmaske wird in eine Binärzahl umgewandelt
        binary = bin(int(okt))[2:]  # dez_in_misc(int(okt), erg=[0])
        subnetzmaske_bin += binary.rjust(8, "0")  # "".join([str(x) for x in bin]).rjust(8, "0")  # "0b" vor Binärzahl wird durch [2:] entfernt; wenn Oktett 0 muss mit 0en aufgefüllt werden
    subnet_range = 2**subnetzmaske_bin.count("0")
    sum_all_hosts = int(np.array(l_geplante_IP_Adressen).sum())
    return sum_all_hosts, subnet_range
    
def geplante_IP_Adressen_minimum():
    ### berechnet die Mindest-Anzahl der IP-Adressen für kleinst-mögliche Subnetze
    l_geplante_IP_Adressen = []
    for host in l_hosts:
        l_geplante_IP_Adressen.append(int(2**np.ceil(math.log2(host))))  # 2**(log2(host) aufgerundet) = die nächst-höhere 2er-Potenz über der Anzahl Hosts
    return l_geplante_IP_Adressen

def subnets_Minimum_Variante():
    """ Dictionary mit geplanten IP-Adressen der Minimum-Variante und Anzahlen der Hosts wird erstellt.
    Wichtig für Übergabe der Werte an andere Funktionen.
    Werte aus anderen Funktionen werden eingeholt und als Argumente an subnets() übergeben, wo Broadcast und Netz-IDs berechnet werden."""
    l_geplante_IP_Adressen = geplante_IP_Adressen_minimum()
    sum_all_hosts, subnet_range = check_genug_IPs(l_geplante_IP_Adressen)
    d_geplante_IP_Adressen = {}
    for name, geplante_IP_Adressen, hosts in zip(l_name, l_geplante_IP_Adressen, l_hosts):
        d_geplante_IP_Adressen[name] = [geplante_IP_Adressen, hosts]
    return subnets(l_geplante_IP_Adressen, sum_all_hosts, subnet_range, d_geplante_IP_Adressen)

def subnets_Verdoppelungs_Variante():
    """ Dictionary mit geplanten IP-Adressen der Verdoppelungs-Variante und Anzahlen der Hosts wird erstellt.
    Wichtig für gemeinsame Sortierung und Übergabe der Werte an andere Funktionen.
    Werte aus anderen Funktionen werden eingeholt und als Argumente an subnets() übergeben, wo Broadcast und Netz-IDs berechnet werden."""
    l_geplante_IP_Adressen = geplante_IP_Adressen_minimum()
    d_geplante_IP_Adressen = {}
    for name, geplante_IP_Adressen, hosts in zip(l_name, l_geplante_IP_Adressen, l_hosts):
        d_geplante_IP_Adressen[name] = [geplante_IP_Adressen, hosts]
    l_freie_Hosts = np.array(l_geplante_IP_Adressen) - np.array(l_hosts)    # Anzahl freier Hosts je Abteilung wird bestimmt
    d = {}
    for name, freie_IPs in zip(l_name, l_freie_Hosts):
        d[name] = freie_IPs     # in einem Dictionary werden die Anzahlen der freien Hosts den Abteilungsnamen zugeordnet
    dsort = dict(sorted(d.items(), key=lambda item: item[1]))   # Dictionary wird sortiert nach Anzahl freier IP-Adressen
    for val in dsort.keys():
        d_geplante_IP_Adressen[val][0] = d_geplante_IP_Adressen[val][0]*2  # geplante IP-Adressen in d_geplante_IP_Adressen werden verdoppelt
        l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())  # Liste mit neuen geplanten IP-Adressen wird erstellt
        sum_all_hosts, subnet_range = check_genug_IPs(list(l_geplante_IP_Adressen))
        if sum_all_hosts > subnet_range:
            d_geplante_IP_Adressen[val][0] = int(d_geplante_IP_Adressen[val][0]/2)  # ist die neue Anzahl zu groß wird sie wieder halbiert
    d_geplante_IP_Adressen = dict(sorted(d_geplante_IP_Adressen.items(), key=lambda item: item[1][0], reverse=True))  # absteigende Sortierung nach geplanten IP-Adressen
    l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())
    return subnets(l_geplante_IP_Adressen, sum_all_hosts, subnet_range, d_geplante_IP_Adressen)
            
def subnets_Prioritaeten_Variante():
    """ Dictionary mit geplanten IP-Adressen der Priorisierungs-Variante und Anzahlen der Hosts wird erstellt.
    Wichtig für gemeinsame Sortierung und Übergabe der Werte an andere Funktionen.
    Werte aus anderen Funktionen werden eingeholt und als Argumente an subnets() übergeben, wo Broadcast und Netz-IDs berechnet werden."""
    l_geplante_IP_Adressen = geplante_IP_Adressen_minimum()
    d_geplante_IP_Adressen = {}
    for i, (name, geplante_IP_Adressen, hosts) in enumerate(zip(l_name, l_geplante_IP_Adressen, l_hosts)):
        if l_prio[i] == "p":
            d_geplante_IP_Adressen[name] = [geplante_IP_Adressen*2, hosts]
        elif l_prio[i] == "pp":
            d_geplante_IP_Adressen[name] = [geplante_IP_Adressen*4, hosts]
        else:
            d_geplante_IP_Adressen[name] = [geplante_IP_Adressen, hosts]
    l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())
    sum_all_hosts, subnet_range = check_genug_IPs(list(l_geplante_IP_Adressen))
    d_geplante_IP_Adressen = dict(sorted(d_geplante_IP_Adressen.items(), key=lambda item: item[1][0], reverse=True))
    l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())
    return subnets(l_geplante_IP_Adressen, sum_all_hosts, subnet_range, d_geplante_IP_Adressen)

def subnets(l_geplante_IP_Adressen, sum_all_hosts, subnet_range, d_geplante_IP_Adressen):
    """ Netz-IDs, Broadcast-Adressen und Subnetzmasken der Subnetze werden berechnet"""
    l_Netz_IDs = [[int(okt) for okt in netz_id_ges.split(".")]]
    l_BC = []
    l_SNM = []
    for i, ip_range in enumerate(l_geplante_IP_Adressen):
        l_BC.append(addierer(l_Netz_IDs[i], [ip_range-1], l=[0]))  # l=[0] muss übergeben werden, da sonst mit der Ergebnisliste des letzten Aufrufs fortgefahren wird
        if i < len(l_geplante_IP_Adressen)-1:  # auf alle Broadcast-Adressen bis auf die letzte wird 1 addiert um die Netz-ID des nächsten Subnetzes zu erhalten
            l_Netz_IDs.append(addierer(l_BC[i], [1], l=[0]))  
    for ip_range in (l_geplante_IP_Adressen):
        ip_range_b256 = dez_in_misc(ip_range, basis=256)
        for i, stelle in enumerate(ip_range_b256):
            if stelle > 0:
                ip_range_b256[i] = 256-stelle
        snm = [255]*(4-len(ip_range_b256)) + ip_range_b256
        l_SNM.append(snm)
    return l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen

def output_excel():
    sheetout.cell(1,1).value = f"Datenüberprüfung abgeschlossen. Es wurden {len(l_fehler)} Fehler gefunden."
    for i, fehler in enumerate(l_fehler):
        sheetout.cell(i+2, 1).value = fehler
        break
    for width in ["H", "J", "L", "V", "X", "Z", "AJ", "AL", "AN"]:
        sheetout.column_dimensions[width].width = 16
    sheetout.cell(3, 2).value = "Netz-Id, Gesamtnetz:"
    sheetout.cell(3, 5).value = netz_id_ges
    sheetout.cell(4, 2).value = "Subnetzmaske, Gesamtnetz:"
    sheetout.cell(4, 5).value = subnetzmaske
    sheetout.cell(6, 2).value = "Gefordert:"
    sheetout.cell(6, 5).value = f"{len(l_hosts)} Subnetze"
    sheetout.cell(8, 2).value = "Von groß nach klein sortiert:"
    count = 0
    for variante in ["Minimum", "Verdoppelung", "Priorisierung"]:
        sheetout.cell(10, 2+count).value = f"Variante {variante}"
        sheetout.cell(10, 2+count).font = Font(bold=True)
        sheetout.cell(11, 8+count).value = "Netz-ID, Subnetz"
        sheetout.cell(11, 8+count).font = Font(bold=True)
        sheetout.cell(11, 10+count).value = "Broadcast, Subnetz"
        sheetout.cell(11, 10+count).font = Font(bold=True)
        sheetout.cell(11, 12+count).value = "Subnetzmaske, Subnetz"
        sheetout.cell(11, 12+count).font = Font(bold=True)
        count += 14

def output_excel_variante(variante):
    if variante == "minimum":
        spalte = 2
    elif variante == "doppelt":
        spalte = 16
    elif variante == "priorisierung":
        spalte = 30
    zeilen = 0
    for i, subnet in enumerate(d_geplante_IP_Adressen.items()):
        sheetout.cell(zeilen+12, spalte).value = f"Subnetz {subnet[0]}: {subnet[1][1]-2} Hosts"
        sheetout.cell(zeilen+12, spalte).font = Font(bold=True)
        sheetout.cell(zeilen+13, spalte).value = "Mindestens benötigte IP-Adressen:"
        sheetout.cell(zeilen+13, spalte+4).alignment = Alignment(horizontal="left")
        sheetout.cell(zeilen+13, spalte+4).value = int(subnet[1][1])
        sheetout.cell(zeilen+14, spalte).value = "Geplante IP-Adressen:"
        sheetout.cell(zeilen+14, spalte+4).alignment = Alignment(horizontal="left")
        sheetout.cell(zeilen+14, spalte+4).value = subnet[1][0]
        sheetout.cell(zeilen+14, spalte+5).value = "von:"
        sheetout.cell(zeilen+14, spalte+6).value = ".".join([str(okt) for okt in l_Netz_IDs[i]])
        sheetout.cell(zeilen+14, spalte+7).value = "bis:"
        sheetout.cell(zeilen+14, spalte+8).value = ".".join([str(okt) for okt in l_BC[i]])
        sheetout.cell(zeilen+14, spalte+10).value = '.'.join([str(okt) for okt in l_SNM[i]])
        zeilen += 5
    sheetout.cell(zeilen+11, spalte).value = "übrige IP-Adressen:"
    sheetout.cell(zeilen+11, spalte+4).alignment = Alignment(horizontal="left")
    l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())
    sheetout.cell(zeilen+11, spalte+4).value = subnet_range - np.array(l_geplante_IP_Adressen).sum()
    
def output_text(variante):
    if variante == "Minimum":
        dateizugriff = "x"
    else:
        dateizugriff = "a"
    with open(freier_name, dateizugriff) as output:
        if variante == "Minimum":
            output.write(f"Datenüberprüfung abgeschlossen. Es wurden {len(l_fehler)} Fehler gefunden.")
            output.write("\n\n")
            output.write(f"{'Netz-ID, Gesamtnetz:':<30}{netz_id_ges}\n")
            output.write(f"{'Subnetzmaske, Gesamtnetz:':<30}{subnetzmaske}")
            output.write("\n\n")
            output.write(f"{'Gefordert:':<30}{len(l_hosts)} Subnetze")
            output.write("\n\n")
            output.write("Von groß nach klein sortiert:")
            output.write("\n\n")
        output.write(f"Variante {variante}\n")
        output.write(f"{'':60}{'Netz-ID, Subnetz':<20}{'':10}{'Broadcast, Subnetz':<20}" +
                        f"{'Subnetzmaske, Subnetz':<20}\n")
        for i, subnet in enumerate(d_geplante_IP_Adressen.items()):
            output.write(f"Subnetz {subnet[0]}: {subnet[1][1]-2} Hosts\n")
            output.write(f"{'Mindestens benötigte IP-Adressen:':<40}{subnet[1][1]}\n")
            output.write(f"{'Geplante IP-Adressen:':<40}{subnet[1][0]:<10}" +
                            f"{'von:':<10}{'.'.join([str(okt) for okt in l_Netz_IDs[i]]):<20}" +
                            f"{'bis:':<10}{'.'.join([str(okt) for okt in l_BC[i]]):<20}" +
                            f"{'.'.join([str(okt) for okt in l_SNM[i]])}\n\n")
        l_geplante_IP_Adressen, l_ = zip(*d_geplante_IP_Adressen.values())
        output.write(f"{'Übrige IP-Adressen:':<40}{subnet_range - np.array(l_geplante_IP_Adressen).sum()}")
        output.write("\n\n\n")



### Programm startet hier!
### Daten werden eingelesen: netz_id_ges = Netz-ID Gesamtnetz, l_hosts = Liste der Hostanzahlen, l_prio = Liste der Prioritäten
check = False
while check == False:
    netz_id_ges, subnetzmaske, l_hosts, l_prio, l_name = daten_einlesen()
    l_fehler = datencheck() ### Daten werden auf Eingabefehler geprüft
    if l_fehler:
        print(f"Datenüberprüfung abgeschlossen. Es wurde(n) {len(l_fehler)} Fehler gefunden.")
        for fehler in l_fehler:
            print(fehler)
        if input("\nBitte korrigieren und neu einlesen.\nJetzt neu einlesen?\n" +
                "(j) = neu einlesen " +
                "(n) = Programm wird beendet ") == "j":
            continue
        else:
            sys.exit()

    ### Erst danach können die Daten in Integers umgewandelt, 2 Hosts für Netz-ID und Broadcastadresse hinzugefügt,
    ### und von groß nach klein sortiert werden
    l_hosts = [int(x)+2 for x in l_hosts]
    l = zip(l_hosts, l_prio, l_name)  #  zusammengehörige Hostanzahl, prio und Abteilungsname werden in tuple gezippt
    ls = sorted(l, reverse=True)  # die Liste der tuple wird absteigend nach Anzahl Hosts sortiert
    l_hosts, l_prio, l_name = zip(*ls)  # tuple-unpacking durch den asterisk-Operator und überschreiben der Listen mit der sortierten Reihenfolge
    l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Minimum_Variante()
    if sum_all_hosts > subnet_range:
        if input(f"{sum_all_hosts} sind zu viele Hosts für ein Gesamtnetz von {subnet_range} IP-Adressen! Neue Daten einlesen?\n" +
                 "(j) = neu einlesen (n) = Programm wird beendet ") == "j":
            continue
        else:
            sys.exit()
    
    try:
        wbout = openpyxl.Workbook()
        sheetout = wbout["Sheet"]
        output_excel()
        output_excel_variante("minimum")
        l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Verdoppelungs_Variante()
        output_excel_variante("doppelt")
        l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Prioritaeten_Variante()
        if sum_all_hosts > subnet_range:
            if input(f"Zu viele Hosts ({sum_all_hosts}) in der Prioritäten-Variante entdeckt. Bei einem Gesamtnetz von {subnet_range} IP-Adressen\n"
                    f"fehlen {sum_all_hosts - subnet_range} IP-Adressen!\n"
                    "Neue Daten einlesen?\n" +
                    "(j) = neu einlesen (n) = Programm wird beendet (Output wird trotzdem erzeugt) ") == "j":
                continue
        output_excel_variante("priorisierung")

        i = 0
        while os.path.exists(f"output{i}.xlsx"):
            i += 1
        wbout.save(f"output{i}.xlsx")

    except:
        i = 0
        while os.path.exists(f"output{i}.txt"):
            i += 1
        freier_name = f"output{i}.txt"
        try:
            l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Minimum_Variante()
            output_text("Minimum") 
            l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Verdoppelungs_Variante()
            output_text("Verdoppelung")
            l_Netz_IDs, l_BC, l_SNM, sum_all_hosts, subnet_range, d_geplante_IP_Adressen = subnets_Prioritaeten_Variante()
            if sum_all_hosts > subnet_range:
                if input(f"Zu viele Hosts ({sum_all_hosts}) in der Prioritäten-Variante entdeckt. Bei einem Gesamtnetz von {subnet_range} IP-Adressen\n"
                    f"fehlen {sum_all_hosts - subnet_range} IP-Adressen!\n"
                    "Neue Daten einlesen?\n" +
                    "(j) = neu einlesen (n) = Programm wird beendet (Output wird trotzdem erzeugt) ") == "j":
                        continue  
            output_text("Priorisierung")   
        except:
            print("Erstellen der Ausgabe-Datei fehlgeschlagen. Programm wird beendet!")
            time.sleep(5)
            sys.exit()
    check = True


